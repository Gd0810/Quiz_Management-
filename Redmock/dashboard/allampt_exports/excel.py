import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
from django.utils import timezone


def generate_attempts_excel(queryset, company, response):
    """
    Generate a professional Excel (.xlsx) report of test attempts.

    Args:
        queryset : Attempt queryset (already filtered)
        company  : Company instance
        response : Django HttpResponse (content_type set to openpyxl type)
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Test Attempts"

    # ---- Styling palette ----------------------------------------------
    C_HEADER_BG   = "0077B6"   # Mid accent blue
    C_HEADER_TEXT = "FFFFFF"   # White text
    C_TITLE_TEXT  = "03045E"   # Deep navy
    C_ZEBRA_BG    = "F4FDFF"   # Light alternate strip tint
    C_BORDER      = "D6E9F5"   # Slightly firmer divider than before (was too faint)
    C_OUTER_BORDER = "0077B6"  # Match header for the table's outer edge

    HEADERS = [
        "Attempt ID", "Candidate Name", "Candidate Email", "Session Type",
        "Test Category", "Difficulty Level", "Total Questions",
        "Correct Answers", "Wrong Answers", "Score (%)", "Date Attempted",
    ]
    NUM_COLS = len(HEADERS)
    TITLE_ROW, SUBTITLE_ROW, HEADER_ROW, FIRST_DATA_ROW = 1, 2, 4, 5

    # ---- Fonts -----------------------------------------------------------
    font_title    = Font(name="Calibri", size=16, bold=True, color=C_TITLE_TEXT)
    font_subtitle = Font(name="Calibri", size=10, italic=True, color="6B7280")
    font_header   = Font(name="Calibri", size=11, bold=True, color=C_HEADER_TEXT)
    font_data     = Font(name="Calibri", size=11, color="1F2937")

    # ---- Alignments --------------------------------------------------------
    align_center = Alignment(horizontal="center", vertical="center")
    align_left   = Alignment(horizontal="left", vertical="center", indent=1)
    align_right  = Alignment(horizontal="right", vertical="center", indent=1)
    align_header = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # ---- Borders -----------------------------------------------------------
    thin_border = Border(
        left=Side(style="thin", color=C_BORDER),
        right=Side(style="thin", color=C_BORDER),
        top=Side(style="thin", color=C_BORDER),
        bottom=Side(style="thin", color=C_BORDER),
    )
    header_border = Border(
        left=Side(style="thin", color=C_HEADER_BG),
        right=Side(style="thin", color=C_HEADER_BG),
        top=Side(style="thin", color=C_HEADER_BG),
        bottom=Side(style="medium", color=C_HEADER_BG),
    )

    # ---- Fills -----------------------------------------------------------
    fill_header = PatternFill(start_color=C_HEADER_BG, end_color=C_HEADER_BG, fill_type="solid")
    fill_zebra  = PatternFill(start_color=C_ZEBRA_BG, end_color=C_ZEBRA_BG, fill_type="solid")

    # ------------------------------------------------------------------
    # Title block (merged across the full table width so it's centered
    # over the data rather than just sitting in column A)
    # ------------------------------------------------------------------
    ws.merge_cells(start_row=TITLE_ROW, start_column=1, end_row=TITLE_ROW, end_column=NUM_COLS)
    title_cell = ws.cell(row=TITLE_ROW, column=1, value=f"{company.name} – Test Attempts Report")
    title_cell.font = font_title
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[TITLE_ROW].height = 28

    ws.merge_cells(start_row=SUBTITLE_ROW, start_column=1, end_row=SUBTITLE_ROW, end_column=NUM_COLS)
    generated_str = timezone.now().strftime("%B %d, %Y  %I:%M %p")
    total_records = queryset.count() if hasattr(queryset, "count") else len(queryset)
    subtitle_cell = ws.cell(
        row=SUBTITLE_ROW, column=1,
        value=f"Generated on {generated_str}   |   {total_records} record{'s' if total_records != 1 else ''}"
    )
    subtitle_cell.font = font_subtitle
    subtitle_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[SUBTITLE_ROW].height = 18
    ws.row_dimensions[3].height = 10  # spacer row

    # ------------------------------------------------------------------
    # Header row
    # ------------------------------------------------------------------
    ws.row_dimensions[HEADER_ROW].height = 28
    for col_idx, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=HEADER_ROW, column=col_idx, value=header)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_header
        cell.border = header_border

    # ------------------------------------------------------------------
    # Data rows
    # ------------------------------------------------------------------
    current_row = FIRST_DATA_ROW
    for attempt in queryset:
        local_created = timezone.localtime(attempt.created_at)
        formatted_date = local_created.strftime("%b %d, %Y, %I:%M %p")

        pct = float(attempt.percentage)
        if pct >= 70:
            score_color = "166534"   # Green
        elif pct >= 50:
            score_color = "92400E"   # Amber
        else:
            score_color = "991B1B"   # Red
        font_score = Font(name="Calibri", size=11, bold=True, color=score_color)

        candidate_name = getattr(attempt.candidate, "name", "") or "—"
        candidate_email = getattr(attempt.candidate, "email", "") or "—"
        level_display = attempt.level.capitalize() if attempt.level else "—"

        row_values = [
            (1, attempt.pk, align_center),
            (2, candidate_name, align_left),
            (3, candidate_email, align_left),
            (4, attempt.get_session_type_display(), align_center),
            (5, attempt.get_test_category_display(), align_left),
            (6, level_display, align_center),
            (7, attempt.question_count, align_right),
            (8, attempt.correct_count, align_right),
            (9, attempt.wrong_count, align_right),
            (11, formatted_date, align_center),
        ]
        for col_idx, value, alignment in row_values:
            cell = ws.cell(row=current_row, column=col_idx, value=value)
            cell.alignment = alignment

        score_cell = ws.cell(row=current_row, column=10, value=pct / 100.0)
        score_cell.number_format = "0.0%"
        score_cell.font = font_score
        score_cell.alignment = align_right

        use_zebra = (current_row % 2 == 0)
        for col_idx in range(1, NUM_COLS + 1):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.border = thin_border
            if col_idx != 10:
                cell.font = font_data
            if use_zebra:
                cell.fill = fill_zebra

        ws.row_dimensions[current_row].height = 20
        current_row += 1

    last_data_row = current_row - 1

    # Friendly empty-state message instead of a table with zero rows
    if last_data_row < FIRST_DATA_ROW:
        ws.merge_cells(start_row=FIRST_DATA_ROW, start_column=1, end_row=FIRST_DATA_ROW, end_column=NUM_COLS)
        empty_cell = ws.cell(row=FIRST_DATA_ROW, column=1, value="No test attempts found for the selected filters.")
        empty_cell.font = Font(name="Calibri", size=11, italic=True, color="6B7280")
        empty_cell.alignment = align_center
        ws.row_dimensions[FIRST_DATA_ROW].height = 24
        last_data_row = FIRST_DATA_ROW

    # ------------------------------------------------------------------
    # Column widths — sized from content, with sane min/max bounds so no
    # single long email or name blows out the whole sheet
    # ------------------------------------------------------------------
    MIN_WIDTH, MAX_WIDTH = 12, 40
    for col_idx in range(1, NUM_COLS + 1):
        col_letter = get_column_letter(col_idx)
        max_len = len(HEADERS[col_idx - 1])
        for row_idx in range(FIRST_DATA_ROW, last_data_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value is None:
                continue
            if cell.number_format == "0.0%" and isinstance(cell.value, float):
                max_len = max(max_len, 6)  # e.g. "100.0%"
            else:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 4, MIN_WIDTH), MAX_WIDTH)

    # ------------------------------------------------------------------
    # Finishing touches for a "proper report" feel
    # ------------------------------------------------------------------
    # Freeze header row + first two ID/name columns so they stay visible
    # while scrolling through a long list.
    ws.freeze_panes = ws.cell(row=FIRST_DATA_ROW, column=1)

    # Autofilter on the header so recipients can sort/filter in Excel.
    if last_data_row >= FIRST_DATA_ROW:
        ws.auto_filter.ref = f"A{HEADER_ROW}:{get_column_letter(NUM_COLS)}{last_data_row}"

    # Clean, borderless grid look (we already draw our own cell borders).
    ws.sheet_view.showGridLines = False

    # Print setup: landscape, fit to one page wide, repeat header row.
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_title_rows = f"{HEADER_ROW}:{HEADER_ROW}"
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5, header=0.2, footer=0.2)

    wb.save(response)