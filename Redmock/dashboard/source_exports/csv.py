from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .common import answer_text, option_rows, subject_name, subtitle_name

content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
extension = 'xlsx'

HEADERS = [
    ('No',                 6),
    ('Subject',           18),
    ('Sub Title',         18),
    ('Level',             10),
    ('Question Paragraph',52),
    ('Question',          52),
    ('Option A',          28),
    ('Option B',          28),
    ('Option C',          28),
    ('Option D',          28),
]

HEADER_FILL  = PatternFill('solid', start_color='1E3A5F')   # deep navy
HEADER_FONT  = Font(name='Poppins', bold=True, size=12, color='FFFFFF')
ROW_FONT     = Font(name='Calibri', size=12, color='1E293B')
ALT_FILL     = PatternFill('solid', start_color='F0F4FA')   # light blue-grey stripe
ANSWER_FILL  = PatternFill('solid', start_color='D1FAE5')   # soft green for answer cols
ANSWER_FONT  = Font(name='Calibri', size=12, color='065F46', bold=True)

THIN = Side(style='thin', color='D1D5DB')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

WRAP = Alignment(wrap_text=True, vertical='top')
CENTER = Alignment(horizontal='center', vertical='top')


def build(quizzes, include_answers=False):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Questions'
    ws.freeze_panes = 'A2'   # freeze header row

    headers = list(HEADERS)
    if include_answers:
        headers += [('Correct Answer', 18), ('Answer Text', 40)]

    # --- Header row ---
    for col_idx, (label, _) in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.border    = BORDER
        cell.alignment = CENTER if label == 'No' else Alignment(
            horizontal='left', vertical='center', wrap_text=False
        )

    ws.row_dimensions[1].height = 22

    # --- Data rows ---
    for index, quiz in enumerate(quizzes, start=1):
        options = [value for _, value in option_rows(quiz)]
        while len(options) < 4:
            options.append('')

        row_data = [
            index,
            subject_name(quiz),
            subtitle_name(quiz),
            quiz.get_level_display(),
            quiz.question_paragraph or '',
            quiz.question,
            options[0],
            options[1],
            options[2],
            options[3],
        ]
        if include_answers:
            row_data += [quiz.correct_answer, answer_text(quiz)]

        excel_row = index + 1
        is_alt    = (index % 2 == 0)

        for col_idx, value in enumerate(row_data, start=1):
            cell           = ws.cell(row=excel_row, column=col_idx, value=value)
            cell.border    = BORDER
            cell.alignment = CENTER if col_idx == 1 else WRAP

            is_answer_col = include_answers and col_idx > len(HEADERS)
            if is_answer_col:
                cell.font = ANSWER_FONT
                cell.fill = ANSWER_FILL
            else:
                cell.font = ROW_FONT
                cell.fill = ALT_FILL if is_alt else PatternFill()

        # auto row height: ~15pt per line, min 18
        max_chars = max(
            len(str(row_data[4] or '')),   # Question Paragraph
            len(str(row_data[5] or '')),   # Question
        )
        col_width = 52  # widest text column width
        lines      = max(1, -(-max_chars // col_width))   # ceiling division
        ws.row_dimensions[excel_row].height = max(18, lines * 15)

    # --- Column widths ---
    for col_idx, (_, width) in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # --- Sheet cosmetics ---
    ws.sheet_view.showGridLines = False

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()